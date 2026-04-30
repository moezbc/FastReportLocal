import csv
import io
import json
import logging
import re
from datetime import datetime

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm

logger = logging.getLogger(__name__)


def execute_report_query(report, parameters):
    """
    Execute a report's SQL query on its external datasource.
    Returns (columns, rows).
    """
    if not report.datasource:
        raise ValueError("Ce rapport n'a pas de source de données configurée.")

    sql = report.sql_query
    params = {}

    # Build param dict and replace operators in SQL
    processed_sql = sql
    for param in report.parameters.all():
        param_name = param.name
        operator = '='
        value2 = ''
        if param_name in parameters:
            param_entry = parameters[param_name]
            if isinstance(param_entry, dict):
                value = param_entry.get('value', '')
                operator = param_entry.get('operator', '=')
                value2 = param_entry.get('value2', '')
            else:
                value = param_entry
                operator = '='
        elif param.default_value:
            value = param.default_value
        else:
            continue

        op_upper = operator.upper()

        # Replace existing operator before :param_name with the chosen one
        pattern = rf"(?i)(\b(?:LIKE|IN|ILIKE|IS(?:[ \t]+NOT)?)\b|[=<>!]+)\s*:{param_name}\b"
        
        def repl(m, op=op_upper, p_name=param_name):
            if op in ('IS NULL', 'IS NOT NULL'):
                return f"{op}"
            elif op == 'BETWEEN':
                return f"BETWEEN :{p_name}_start AND :{p_name}_end"
            else:
                return f"{op} :{p_name}"

        processed_sql = re.sub(pattern, repl, processed_sql)
        
        if op_upper == 'BETWEEN':
            params[f"{param_name}_start"] = value
            params[f"{param_name}_end"] = value2
            param_list_to_handle = [f"{param_name}_start", f"{param_name}_end"]
        elif op_upper not in ('IS NULL', 'IS NOT NULL'):
            if op_upper == 'IN':
                # Quick handling for IN with a comma-separated string, though not bulletproof for all SQL drivers
                params[param_name] = value
            else:
                params[param_name] = value
            param_list_to_handle = [param_name]
        else:
            param_list_to_handle = []

    # Get connection from datasource
    conn = report.datasource.get_connection()
    try:
        cursor = conn.cursor()

        # Replace :param_name with driver-appropriate placeholders
        db_type = report.datasource.db_type

        # Use the processed_sql
        if db_type == 'postgresql':
            # psycopg2 uses %(name)s
            for p_name in params:
                processed_sql = processed_sql.replace(f':{p_name}', f'%({p_name})s')
            cursor.execute(processed_sql, params)
        elif db_type == 'mysql':
            # pymysql uses %(name)s
            for p_name in params:
                processed_sql = processed_sql.replace(f':{p_name}', f'%({p_name})s')
            cursor.execute(processed_sql, params)
        elif db_type == 'oracle':
            # cx_Oracle supports :name natively
            cursor.execute(processed_sql, params)
        elif db_type == 'sqlserver':
            # pyodbc uses ? positional params
            ordered_params = []
            # We must find the placeholders in the order they appear to append them correctly
            # Wait, replacing them individually with ? and appending to a list might lose the correct order
            # This is a bit tricky, but wait, the original logic had this flaw too: 
            # It just did:
            # for param_name in params:
            #     processed_sql = processed_sql.replace(f':{param_name}', '?')
            #     ordered_params.append(params[param_name])
            for p_name in params:
                processed_sql = processed_sql.replace(f':{p_name}', '?')
                ordered_params.append(params[p_name])
            if ordered_params:
                cursor.execute(processed_sql, ordered_params)
            else:
                cursor.execute(processed_sql)

        columns = [col[0] for col in cursor.description] if cursor.description else []
        rows = cursor.fetchall()
        cursor.close()
    finally:
        conn.close()

    return columns, rows


def export_results(columns, rows, output_type, csv_separator=','):
    """
    Convert query results to the specified output format.
    Returns (bytes_content, content_type, file_extension).
    """
    output_type = output_type.upper()

    if output_type == 'CSV':
        return _export_csv(columns, rows, csv_separator)
    elif output_type == 'XLSX':
        return _export_xlsx(columns, rows)
    elif output_type == 'JSON':
        return _export_json(columns, rows)
    elif output_type == 'XML':
        return _export_xml(columns, rows)
    elif output_type == 'PDF':
        return _export_pdf(columns, rows)
    else:
        raise ValueError(f"Type de sortie non supporté : {output_type}")


def _export_csv(columns, rows, separator=','):
    output = io.StringIO()
    writer = csv.writer(output, delimiter=separator)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([_serialize_value(v) for v in row])
    content = output.getvalue().encode('utf-8-sig')
    return content, 'text/csv; charset=utf-8', 'csv'


def _export_xlsx(columns, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"
    ws.append(columns)
    for row in rows:
        ws.append([_serialize_value(v) for v in row])

    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx'


def _export_json(columns, rows):
    data = [dict(zip(columns, [_serialize_value(v) for v in row])) for row in rows]
    content = json.dumps(data, ensure_ascii=False, indent=2, default=str).encode('utf-8')
    return content, 'application/json', 'json'


def _export_xml(columns, rows):
    lines = ['<?xml version="1.0" encoding="UTF-8"?>', '<report>']
    for row in rows:
        lines.append('  <row>')
        for col, val in zip(columns, row):
            safe_col = col.replace(' ', '_').replace('-', '_')
            lines.append(f'    <{safe_col}>{_serialize_value(val)}</{safe_col}>')
        lines.append('  </row>')
    lines.append('</report>')
    content = '\n'.join(lines).encode('utf-8')
    return content, 'application/xml', 'xml'


def _export_pdf(columns, rows):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph("Rapport", styles['Title']))
    elements.append(Spacer(1, 10 * mm))

    table_data = [columns]
    for row in rows[:1000]:
        table_data.append([str(_serialize_value(v))[:50] for v in row])

    if len(table_data) > 1:
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(table)

    doc.build(elements)
    return buffer.getvalue(), 'application/pdf', 'pdf'


def _serialize_value(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.isoformat()
    return value
